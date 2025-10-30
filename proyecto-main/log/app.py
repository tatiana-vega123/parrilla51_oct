print("🔥 Cargando archivo app.py desde ruta activa")

# app.py unificado (comentado por bloques y funciones)
from flask import (
    Flask, render_template, request, redirect, url_for, session,
    jsonify, flash, send_file
)
from flask_mysqldb import MySQL
from flask_mail import Mail   # falta message
# from werkzeug.security import generate_password_hash, check_password_hash


# from functools import wraps
# import uuid
import io
from MySQLdb.cursors import DictCursor
# para la fecha de la reserva
from datetime import datetime
# Reportes
from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl


# ------------------ CONFIG ------------------
# Aquí se configura la app Flask, la conexión a MySQL y el servicio de correo.
# También se define la secret_key para sesiones.
app = Flask(__name__, template_folder="templates")

app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'parrilla51'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'

mysql = MySQL(app)

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'andresfariasa@juandelcorral.edu.co'
app.config['MAIL_PASSWORD'] = 'ekyxnfjsubefudgm'
mail = Mail(app)

app.secret_key = "pinchellave"

# ------------------ DECORADOR LOGIN ------------------
# Decorador para proteger rutas que requieren autenticación.
# Si no hay sesión con 'logueado', redirige al login y muestra aviso.

# ------------------ AUTENTICACIÓN (ruta raíz) ------------------
# La misma vista maneja '/' y '/login' (registrada también como
# endpoint 'login').
# - GET: muestra template index.html
# - POST: procesa login (busca usuario activo y valida contraseña)


# {@app.route('/', methods=['GET', 'POST']) #}

# def index():
# Manejo login si se hace POST (compatible con formularios
# que hagan action a '/')
# if request.method == 'POST':
# correo = request.form.get('txtCorreo')
# password = request.form.get('txtPassword')
# cur = mysql.connection.cursor()
# cur.execute('SELECT * FROM usuarios WHERE correo=%s '
# 'AND estado="activo"', (correo,))
# account = cur.fetchone()
# if account and check_password_hash(account.get('contraseña', ''),
# password):
# Crear sesión y guardar datos importantes
# session['logueado'] = True
# session['id_usuario'] = account['id_usuario']
# session['nombre'] = account['nombre']
# session['rol'] = account.get('rol', 'cliente')
# redirigir según rol
# if session['rol'] == 'administrador':
# return redirect(url_for("admin_dashboard"))
# elif session['rol'] == 'cliente':
# return redirect(url_for("cliente_dashboard"))
# elif session['rol'] == 'empleado':
# return redirect(url_for("empleado_dashboard"))
# else:
# Mensaje sencillo si falla autenticación o cuenta no está activa
# return render_template('index.html', mensaje="Usuario o contraseña"
#  "incorrectos, o cuenta no activada")
# return render_template('index.html')

# Registrar alias '/login' con endpoint 'login' apuntando a la misma vista
# index


# app.add_url_rule('/login', endpoint='login', view_func=index,
#               methods=['GET', 'POST'])

# Ruta para mostrar formulario de registro


# { @app.route('/registro')#}
# def registro():
# return render_template('registro.html')

# Guardar usuario: recibe datos del formulario, encripta contraseña, crea token
# y envía correo de activación con enlace que apunta a /activar/<token>


# {@app.route('/guardar-usuario', methods=["POST"])#}
# def guardar_usuario():
# nombre = request.form['nombre']
# apellido = request.form['apellido']
# telefono = request.form['telefono']
# direccion = request.form['direccion']
# correo = request.form['correo']
# password = request.form['password']
# rol = request.form.get('rol', 'cliente')

# password_hash = generate_password_hash(password)
# token = str(uuid.uuid4())

# cur = mysql.connection.cursor()
# cur.execute("""
# INSERT INTO usuarios(nombre, apellido, telefono, direccion, correo,
# contraseña, rol, estado, token_activacion)
# VALUES(%s,%s,%s,%s,%s,%s,%s,'inactivo',%s)
# """, (nombre, apellido, telefono, direccion, correo,
# password_hash, rol, token))
# mysql.connection.commit()

# enlace = url_for('activar_cuenta', token=token, _external=True)
# try:
# msg = Message('Activa tu cuenta', sender=app.config['MAIL_USERNAME'],
# recipients=[correo])
# msg.body = f'Hola {nombre}, haz clic en el siguiente enlace para'
# f'activar tu cuenta:\n{enlace}'
# mail.send(msg)
# except Exception as e:
# Loguea error de envío de correo en consola; no romper la experiencia
# de registro
# print("Error enviando correo:", e)

# return render_template('index.html',
# mensaje="Revisa tu correo para activar tu cuenta")

# Activación de cuenta: recibe token, busca usuario y cambia estado a 'activo'


# {@app.route('/activar/<token>')#}
# def activar_cuenta(token):
# cur = mysql.connection.cursor()
# cur.execute("SELECT * FROM usuarios WHERE token_activacion=%s", (token,))
# usuario = cur.fetchone()
# if usuario:
# cur.execute("UPDATE usuarios SET estado='activo'",
# "token_activacion=NULL WHERE id_usuario=%s",
# (usuario['id_usuario'],))
# mysql.connection.commit()
# return ("Tu cuenta ha sido activada correctamente",
# "Ya puedes iniciar sesión.")

# else:
# return "Token inválido o expirado."

# Ruta histórica para compatibilidad con templates antiguos
# Redirige al index que ya maneja el POST de login


# @app.route('/acceso-login', methods=["POST"])
# def acceso_login():
#    return index()

# ------------------ DASHBOARDS ------------------
# Vistas protegidas por login_requerido. Cada una muestra el template
# correspondiente.


@app.route("/admin")
def admin_dashboard():
    return render_template("admin2.html")


@app.route("/cliente/dashboard")
def cliente_dashboard():
    return render_template("cliente_dashboard.html")


@app.route("/empleado")
def empleado_dashboard():
    return render_template("empleado.html")

# ------------------ PERFIL ------------------
# GET: devuelve datos del perfil del usuario logueado en formato JSON
# POST: actualiza el perfil del usuario (recibe JSON con nuevos datos)


@app.route('/perfil', methods=['GET'])
def perfil():
    cur = mysql.connection.cursor()
    cur.execute("SELECT nombre, apellido, telefono, direccion,"
                "correo FROM usuarios WHERE id_usuario=%s",
                (session['id_usuario'],))
    usuario = cur.fetchone()
    return jsonify(usuario or {})


@app.route('/perfil', methods=['POST'])
def actualizar_perfil():
    data = request.json
    cur = mysql.connection.cursor()
    cur.execute("""
        UPDATE usuarios
        SET nombre=%s, apellido=%s, telefono=%s, direccion=%s, correo=%s
        WHERE id_usuario=%s
    """, (data['nombre'], data['apellido'], data['telefono'],
          data['direccion'],
          data['correo'], session['id_usuario']))
    mysql.connection.commit()
    return jsonify({'mensaje': 'Perfil actualizado correctamente'})

# ------------------ BLUEPRINT-LIKE ROUTES (integradas) ------------------
# Implementación de rutas que actúan como si fueran blueprints
# integrados en un único archivo.
# Se crean alias para endpoints que templates antiguos
# podrían usar (evita errores con url_for).
# Reservas (vista sencilla que lista todas las reservas)


@app.route("/cliente_reservar", methods=["GET", "POST"])
def cliente_reservar():
    cur = mysql.connection.cursor()

    if request.method == "POST":
        fecha = request.form["fecha"]
        hora = request.form["hora"]
        cant_personas = request.form["cant_personas"]
        telefono = request.form["telefono"]

        # Guardar en MySQL (sin id_mesa)
        cur.execute("""
            INSERT INTO reservas (fecha, hora, cant_personas, telefono)
            VALUES (%s, %s, %s, %s)
        """, (fecha, hora, cant_personas, telefono))
        mysql.connection.commit()

        flash("✅ Reserva creada con éxito", "success")
        return redirect(url_for("ver_reservas"))

    # Si es GET: ya no necesitas cargar mesas
    return render_template("cliente_reservar.html")


# ------------------ FUNCIONES AUX: CATEGORIAS ------------------
# Función auxiliar para obtener sólo ciertas categorías importantes


def obtener_categorias():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT * FROM categorias
        WHERE nombre_categoria IN ('Bebidas', 'Res', 'Pollo',
        'Cerdo', 'Plato del dia',
        'Entradas', 'Acompañamientos', 'Platos combinados',
        'Cortes gruesos','Adicionales')
    """)
    return cur.fetchall()

# ------------------ INVENTARIO y CRUD ------------------
# Vista inicial admin


@app.route("/inicioadmin")
def inicioadmin():
    return render_template("inicioadmin.html")

# Inventario: reúne productos, insumos, mesas y subcategorías para la vista


@app.route("/inventario")
def inventario():
    cur = mysql.connection.cursor()
    # Productos con categoría
    # (LEFT JOIN permite que no haya categoría sin fallar)
    cur.execute("""
        SELECT p.id_producto, p.nombre, p.cantidad, p.descripcion, p.precio,
               p.imagen, c.nombre_categoria, p.cod_categoria
        FROM productos p
        LEFT JOIN categorias c ON p.cod_categoria = c.id_categoria
    """)
    productos = cur.fetchall()

    # Insumos con su subcategoría
    cur.execute("""
        SELECT i.id_insumo, i.nombre, i.cantidad, i.precio,
        i.fecha_vencimiento, i.lote,
               s.nombre_subcategoria, i.subcategoria_id
        FROM insumos i
        LEFT JOIN subcategorias_insumos s ON
        i.subcategoria_id = s.id_subcategoria
    """)
    insumos = cur.fetchall()

    # Mesas
    cur.execute("SELECT * FROM mesas")
    mesas = cur.fetchall()

    categorias = obtener_categorias()
    cur.execute("SELECT * FROM subcategorias_insumos")
    subcategorias = cur.fetchall()

    return render_template("inventario.html",
                           productos=productos,
                           insumos=insumos,
                           mesas=mesas,
                           categorias=categorias,
                           subcategorias=subcategorias)

# CRUD Productos: agregar producto
# (GET muestra formulario, POST procesa y guarda)


@app.route("/producto/agregar", methods=["GET", "POST"])
def agregar_producto():
    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            cantidad = int(request.form["cantidad"])
            descripcion = request.form["descripcion"].strip()
            # Una sola línea, limpio y correcto
            precio = int(request.form
                         ["precio"].replace(".", "").replace(",", ""))

            cod_categoria = int(request.form["cod_categoria"])
            imagen = request.form.get("imagen", "").strip()
        except (ValueError, KeyError):
            return "Datos inválidos", 400

        cur = mysql.connection.cursor()
        cur.execute("""
            INSERT INTO productos (nombre, cantidad, descripcion, precio,
            cod_categoria, imagen)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (nombre, cantidad, descripcion, precio, cod_categoria, imagen))
        mysql.connection.commit()
        return redirect(url_for("inventario"))

    categorias = obtener_categorias()
    return render_template("editar_producto.html",
                           producto=None, categorias=categorias)

# Editar producto: GET carga producto; POST actualiza


@app.route("/producto/editar/<int:id_producto>", methods=["GET", "POST"])
def editar_producto(id_producto):
    cur = mysql.connection.cursor()
    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            cantidad = int(request.form["cantidad"])
            descripcion = request.form["descripcion"].strip()
            precio = int(request.form
                         ["precio"].replace(".", "").replace(",", ""))
            cod_categoria = int(request.form["cod_categoria"])
            imagen = request.form.get("imagen", "").strip()
        except (ValueError, KeyError):
            return "Datos inválidos", 400

        cur.execute("""
            UPDATE productos
            SET nombre=%s, cantidad=%s, descripcion=%s, precio=%s,
            cod_categoria=%s, imagen=%s
            WHERE id_producto=%s
        """, (nombre, cantidad, descripcion, precio, cod_categoria, imagen,
              id_producto))
        mysql.connection.commit()
        return redirect(url_for("inventario"))

    cur.execute("SELECT * FROM productos WHERE id_producto=%s", (id_producto,))
    producto = cur.fetchone()
    categorias = obtener_categorias()
    if not producto:
        return "Producto no encontrado", 404
    return render_template("editar_producto.html", producto=producto,
                           categorias=categorias)

# Eliminar producto (POST para evitar CSRF por GET)


@app.route("/producto/eliminar/<int:id_producto>", methods=["POST"])
def eliminar_producto(id_producto):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM productos WHERE id_producto=%s", (id_producto,))
    mysql.connection.commit()
    return redirect(url_for("inventario"))

# CRUD Insumos: estructuras paralelas a productos (agregar, editar, eliminar)


@app.route("/insumo/agregar", methods=["GET", "POST"])
def agregar_insumo():
    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            cantidad = int(request.form["cantidad"])
            precio = float(request.form["precio"])
            fecha_vencimiento = request.form.get("fecha_vencimiento") or None
            lote = request.form.get("lote") or None
            subcategoria_id = int(request.form["subcategoria_id"])
        except (ValueError, KeyError):
            return "Datos inválidos", 400

        cur = mysql.connection.cursor()
        cur.execute("""
            INSERT INTO insumos (nombre, cantidad, precio, fecha_vencimiento,
            lote, subcategoria_id)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (nombre, cantidad, precio, fecha_vencimiento, lote,
              subcategoria_id))
        mysql.connection.commit()
        return redirect(url_for("inventario"))

    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM subcategorias_insumos")
    subcategorias = cur.fetchall()
    return render_template("editar_insumo.html", insumo=None,
                           subcategorias=subcategorias)


@app.route("/insumo/editar/<int:id_insumo>", methods=["GET", "POST"])
def editar_insumo(id_insumo):
    cur = mysql.connection.cursor()
    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            cantidad = int(request.form["cantidad"])
            precio = float(request.form["precio"])
            fecha_vencimiento = request.form.get("fecha_vencimiento") or None
            lote = request.form.get("lote") or None
            subcategoria_id = int(request.form["subcategoria_id"])
        except (ValueError, KeyError):
            return "Datos inválidos", 400

        cur.execute("""
            UPDATE insumos
            SET nombre=%s, cantidad=%s, precio=%s, fecha_vencimiento=%s,
            lote=%s, subcategoria_id=%s
            WHERE id_insumo=%s
        """, (nombre, cantidad, precio, fecha_vencimiento, lote,
              subcategoria_id, id_insumo))
        mysql.connection.commit()
        return redirect(url_for("inventario"))

    cur.execute("SELECT * FROM insumos WHERE id_insumo=%s", (id_insumo,))
    insumo = cur.fetchone()
    cur.execute("SELECT * FROM subcategorias_insumos")
    subcategorias = cur.fetchall()
    if not insumo:
        return "Insumo no encontrado", 404
    return render_template("editar_insumo.html", insumo=insumo,
                           subcategorias=subcategorias)


@app.route("/insumo/eliminar/<int:id_insumo>", methods=["POST"])
def eliminar_insumo(id_insumo):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM insumos WHERE id_insumo=%s", (id_insumo,))
    mysql.connection.commit()
    return redirect(url_for("inventario"))

# ------------------ CRUD Mesas ------------------
# Agregar mesa, cambiar estado y eliminar


@app.route("/mesa/agregar", methods=["GET", "POST"])
def agregar_mesa():
    if request.method == "POST":
        try:
            numero = int(request.form["numero"])
            capacidad = int(request.form["capacidad"])
        except (ValueError, KeyError):
            return "Datos inválidos", 400
        cur = mysql.connection.cursor()
        cur.execute("""
            INSERT INTO mesas (numero, capacidad, estado)
            VALUES (%s, %s, 'Disponible')
        """, (numero, capacidad))
        mysql.connection.commit()
        return redirect(url_for("inventario"))
    return render_template("editar_mesa.html", mesa=None)


@app.route("/mesa/cambiar_estado/<int:id_mesa>")
def cambiar_estado(id_mesa):
    cur = mysql.connection.cursor()
    cur.execute("SELECT estado FROM mesas WHERE id_mesa=%s", (id_mesa,))
    mesa = cur.fetchone()
    if not mesa:
        return "Mesa no encontrada", 404
    nuevo_estado = (
        "Disponible" if mesa["estado"] == "No disponible" else "No disponible"
        )
    cur.execute("UPDATE mesas SET estado=%s WHERE id_mesa=%s",
                (nuevo_estado, id_mesa))
    mysql.connection.commit()
    return redirect(url_for("inventario"))


@app.route("/mesa/eliminar/<int:id_mesa>", methods=["POST"])
def eliminar_mesa(id_mesa):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM mesas WHERE id_mesa=%s", (id_mesa,))
    mysql.connection.commit()
    return redirect(url_for("inventario"))

# Confirmación de eliminación: prepara URL para el POST real de eliminación


@app.route("/confirmar_eliminacion/<string:tipo>/<int:item_id>")
def confirmar_eliminacion(tipo, item_id):
    if tipo == "producto":
        url_elim = url_for("eliminar_producto", id_producto=item_id)
    elif tipo == "insumo":
        url_elim = url_for("eliminar_insumo", id_insumo=item_id)
    elif tipo == "mesa":
        url_elim = url_for("eliminar_mesa", id_mesa=item_id)
    else:
        return "Tipo no válido", 400
    return render_template("eliminar.html", tipo=tipo,
                           item_id=item_id, url=url_elim)

# ------------------ ASIGNAR ROL / USUARIOS ------------------
# Vista para listar usuarios y asignar rol/estado via formulario POST


@app.route("/asignarol", methods=["GET", "POST"])
def asignarol():
    if request.method == "POST":
        user_id = request.form["id_usuario"]
        nuevo_rol = request.form["rol"]
        nuevo_estado = request.form["estado"]
        cur = mysql.connection.cursor()
        cur.execute("""
            UPDATE usuarios
            SET rol = %s, estado = %s
            WHERE id_usuario = %s
        """, (nuevo_rol, nuevo_estado, user_id))
        mysql.connection.commit()
        return redirect(url_for("asignarol"))
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM usuarios")
    usuarios = cur.fetchall()
    return render_template("asignarol.html", usuarios=usuarios)

# Rutas helper para cambiar estado/rol vía URL
# (útiles para botones rápidos en la UI)


@app.route("/usuario/cambiar_estado/<int:id_usuario>/<string:nuevo_estado>")
def cambiar_estado_usuario(id_usuario, nuevo_estado):
    cur = mysql.connection.cursor()
    cur.execute("UPDATE usuarios SET estado = %s WHERE id_usuario = %s",
                (nuevo_estado, id_usuario))
    mysql.connection.commit()
    return redirect(url_for("asignarol"))


@app.route("/usuario/cambiar_rol/<int:id_usuario>/<string:nuevo_rol>")
def cambiar_rol_usuario(id_usuario, nuevo_rol):
    cur = mysql.connection.cursor()
    cur.execute("UPDATE usuarios SET rol = %s WHERE id_usuario = %s",
                (nuevo_rol, id_usuario))
    mysql.connection.commit()
    return redirect(url_for("asignarol"))

# ------------------ CONSULTAS Y REPORTES ------------------
# Consultas específicas para reservas y ventas, con filtros por estado/tipo


@app.route("/consultar_reservas")
def consultar_reservas():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT id_reserva, fecha, hora, cant_personas, estado, cod_mesa,
        telefono, id_usuario
        FROM reservas
        WHERE estado IN ('aceptada', 'cancelada')
        ORDER BY fecha ASC, hora ASC
    """)
    reservas = cur.fetchall()
    return render_template("consultar_reservas.html", reservas=reservas)


@app.route("/consultaVentas")
def consultaVentas():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora,
        p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'restaurante' AND p.estado IN ('entregado',
        'cancelado')
        ORDER BY p.fecha DESC, p.hora DESC
    """)
    pedidos_restaurante = cur.fetchall()

    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha,
        p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'domicilio' AND p.estado IN ('entregado',
        'cancelado')
        ORDER BY p.fecha DESC, p.hora DESC
    """)
    pedidos_domicilio = cur.fetchall()

    return render_template("consultaVentas.html",
                           pedidos_restaurante=pedidos_restaurante,
                           pedidos_domicilio=pedidos_domicilio)

# Consultas de inventario para vistas específicas


@app.route("/consulta_P")
def consulta_P():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_producto, p.nombre, p.cantidad, p.descripcion, p.precio,
               p.imagen, c.nombre_categoria
        FROM productos p
        LEFT JOIN categorias c ON p.cod_categoria = c.id_categoria
        ORDER BY p.nombre ASC
    """)
    productos = cur.fetchall()
    return render_template("consulta_P.html", productos=productos)


@app.route("/consulta_Y")
def consulta_Y():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT i.id_insumo, i.nombre, i.cantidad, i.precio,
        i.fecha_vencimiento, i.lote,
               s.nombre_subcategoria
        FROM insumos i
        LEFT JOIN subcategorias_insumos s ON
        i.subcategoria_id = s.id_subcategoria
        ORDER BY i.nombre ASC
    """)
    insumos = cur.fetchall()
    return render_template("consulta_Y.html", insumos=insumos)

# ------------------ REPORTES (pantalla / pdf / excel) ------------------
# Vista que consolida distintos datos para el módulo de reportes


@app.route("/reportes")
def reportes():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha,
        p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'restaurante'
    """)
    pedidos_restaurante = cur.fetchall()

    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido,
        p.fecha, p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'domicilio'
    """)
    pedidos_domicilio = cur.fetchall()

    cur.execute("SELECT * FROM vista_reservas_mesas")
    reservas = cur.fetchall() or []

    cur.execute("SELECT * FROM vista_insumos_stock_bajo")
    stock_bajo = cur.fetchall() or []

    return render_template("reportes.html",
                           pedidos_restaurante=pedidos_restaurante or [],
                           pedidos_domicilio=pedidos_domicilio or [],
                           reservas=reservas,
                           stock_bajo=stock_bajo)

# Genera PDF con reportes: pedidos restaurante, pedidos domicilio,
# reservas e inventario bajo


@app.route("/reportes/pdf")
def reportes_pdf():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha,
        p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'restaurante'
    """)
    pedidos_restaurante = cur.fetchall()

    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido,
        p.fecha, p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'domicilio'
    """)
    pedidos_domicilio = cur.fetchall()

    cur.execute("SELECT * FROM vista_reservas_mesas")
    reservas = cur.fetchall()

    cur.execute("SELECT * FROM insumos WHERE cantidad < 5")
    inventario = cur.fetchall()

    # Construcción de PDF usando reportlab
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Pedidos restaurante
    elements.append(Paragraph("Pedidos en Restaurante", styles['Heading2']))
    data = [["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"]]
    if pedidos_restaurante:
        for p in pedidos_restaurante:
            data.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}",
                         str(p['fecha']), str(p['hora']),
                         f"${p['total']:.2f}", p['estado']])
    else:
        data.append(["-", "No hay pedidos en restaurante", "-", "-", "-", "-"])
    elements.append(Table(data))
    elements.append(Spacer(1, 12))

    # Pedidos domicilio
    elements.append(Paragraph("Pedidos a Domicilio", styles['Heading2']))
    data = [["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"]]
    if pedidos_domicilio:
        for p in pedidos_domicilio:
            data.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}",
                         str(p['fecha']), str(p['hora']),
                         f"${p['total']:.2f}", p['estado']])
    else:
        data.append(["-", "No hay pedidos a domicilio", "-", "-", "-", "-"])
    elements.append(Table(data))
    elements.append(Spacer(1, 12))

    # Reservas
    elements.append(Paragraph("Reservas", styles['Heading2']))
    data = [["ID", "Fecha", "Hora", "Personas", "Estado", "Mesa", "Capacidad"]]
    if reservas:
        for r in reservas:
            data.append([r['id_reserva'], str(r['fecha']), str(r['hora']),
                         r['cant_personas'], r['estado'], r.get('mesa'),
                         r.get('capacidad')])
    else:
        data.append(["-", "No hay reservas registradas", "-", "-",
                     "-", "-", "-"])
    elements.append(Table(data))
    elements.append(Spacer(1, 12))

    # Inventario bajo
    elements.append(Paragraph("Inventario Bajo", styles['Heading2']))
    data = [["ID Insumo", "Nombre", "Cantidad", "Precio"]]
    if inventario:
        for i in inventario:
            data.append([i['id_insumo'], i['nombre'], i['cantidad'],
                         f"${i['precio']:.2f}"])
    else:
        data.append(["-", "No hay insumos con stock bajo", "-", "-"])
    elements.append(Table(data))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name="reporte_general.pdf",
                     mimetype="application/pdf")

# Genera Excel con reportes análogos al PDF


@app.route("/reportes/excel")
def reportes_excel():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha,
        p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'restaurante'
    """)
    pedidos_restaurante = cur.fetchall()

    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido,
        p.fecha, p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'domicilio'
    """)
    pedidos_domicilio = cur.fetchall()

    cur.execute("SELECT * FROM vista_reservas_mesas")
    reservas = cur.fetchall()

    cur.execute("SELECT * FROM insumos WHERE cantidad < 5")
    inventario = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Restaurante"
    ws.append(["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"])
    if pedidos_restaurante:
        for p in pedidos_restaurante:
            ws.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}",
                       str(p['fecha']), str(p['hora']),
                       p['total'], p['estado']])
    else:
        ws.append(["-", "No hay pedidos en restaurante", "-", "-", "-", "-"])

    ws = wb.create_sheet("Pedidos Domicilio")
    ws.append(["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"])
    if pedidos_domicilio:
        for p in pedidos_domicilio:
            ws.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}",
                       str(p['fecha']), str(p['hora']),
                       p['total'], p['estado']])
    else:
        ws.append(["-", "No hay pedidos a domicilio", "-", "-", "-", "-"])

    ws = wb.create_sheet("Reservas")
    ws.append(["ID Reserva", "Fecha", "Hora", "Personas",
               "Estado", "Mesa", "Capacidad"])
    if reservas:
        for r in reservas:
            ws.append([r['id_reserva'], str(r['fecha']),
                       str(r['hora']), r['cant_personas'], r['estado'],
                       r.get('mesa'), r.get('capacidad')])
    else:
        ws.append(["-", "No hay reservas registradas", "-",
                   "-", "-", "-", "-"])

    ws = wb.create_sheet("Inventario Bajo")
    ws.append(["ID Insumo", "Nombre", "Cantidad", "Precio"])
    if inventario:
        for i in inventario:
            ws.append([i['id_insumo'], i['nombre'],
                       i['cantidad'], i['precio']])
    else:
        ws.append(["-", "No hay insumos con stock bajo", "-", "-"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer, as_attachment=True,
        download_name="reporte_general.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument."
                 "spreadsheetml.sheet")

# ------------------ ADMIN: PEDIDOS ------------------
# Vista de pedidos para administrador y función para cambiar estado


@app.route("/admin/pedidos")
def admin_pedidos():
    if session.get("rol") != "administrador":
        flash("Acceso denegado", "danger")
        return redirect(url_for("index"))

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora,
        p.total, p.estado, p.tipo_entrega, p.cod_mesa
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        ORDER BY p.fecha DESC, p.hora DESC
    """)
    pedidos = cur.fetchall()

    return render_template("admin_pedidos.html", pedidos=pedidos)


@app.route(
    "/admin/pedidos/cambiar_estado/<int:id_pedido>/<string:nuevo_estado>"
    )
def cambiar_estado_pedido(id_pedido, nuevo_estado):
    if session.get("rol") != "administrador":
        flash("Acceso denegado", "danger")
        return redirect(url_for("index"))

    cur = mysql.connection.cursor()
    cur.execute("UPDATE pedidos SET estado=%s WHERE id_pedido=%s",
                (nuevo_estado, id_pedido))
    mysql.connection.commit()
    flash(f"Estado del pedido {id_pedido} cambiado a {nuevo_estado}",
          "success")
    return redirect(url_for("admin_pedidos"))

# ------------------ RUTAS CLIENTE ------------------
# Cliente puede reservar, ver productos y usar carrito/pedidos


@app.route("/cliente/reservar/form", methods=["GET", "POST"])
def cliente_reservar_form():
    if request.method == "POST":
        nombre = request.form["nombre"]
        documento = request.form["documento"]
        fecha = request.form["fecha"]
        hora = request.form["hora"]
        cant_personas = int(request.form["cant_personas"])
        tipo_evento = request.form["tipo_evento"]
        comentarios = request.form["comentarios"]
        telefono = request.form["telefono"]

        # Validar fecha y hora
        fecha_reserva = datetime.strptime(fecha, "%Y-%m-%d").date()
        hoy = datetime.now().date()

        if fecha_reserva < hoy:
            flash("❌ No puedes reservar en una fecha pasada.", "error")
            return redirect(url_for("cliente_reservar_form"))

        if fecha_reserva == hoy:
            hora_actual = datetime.now().time()
            hora_reserva = datetime.strptime(hora, "%H:%M").time()
            if hora_reserva < hora_actual:
                flash("❌ No puedes reservar en una hora pasada.", "error")
                return redirect(url_for("cliente_reservar_form"))

        #  Guardar la reserva en la base de datos
        cur = mysql.connection.cursor()
        cur.execute("""
            INSERT INTO reservas (nombre, documento, fecha, hora,
            cant_personas, tipo_evento, comentarios, estado, telefono,
            id_usuario)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'pendiente', %s, %s)
        """, (
            nombre,
            documento,
            fecha,
            hora,
            cant_personas,
            tipo_evento,
            comentarios,
            telefono,
            1
        ))

        mysql.connection.commit()
        flash(" Reserva registrada correctamente. Espera confirmación.",
              "success")
        return redirect(url_for("cliente_dashboard"))

    return render_template("cliente_reservar.html")


@app.route('/cancelar/<int:id>', methods=['POST'])
def cancelar_reserva(id):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM reservas WHERE id_reserva = %s", (id,))
    mysql.connection.commit()
    cur.close()
    flash("Reserva cancelada exitosamente", "success")
    return redirect(url_for('ver_reservas'))


@app.route("/mis-reservas", methods=["GET"])
def ver_reservas():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM reservas")
    reservas = cur.fetchall()
    return render_template("cliente_ver_reservas.html", reservas=reservas)


# Mostrar productos disponibles al cliente

@app.route("/cliente/productos")
def cliente_productos():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_producto, p.nombre, p.descripcion, p.precio,
        p.imagen, c.nombre_categoria
        FROM productos p
        LEFT JOIN categorias c ON p.cod_categoria = c.id_categoria
        ORDER BY p.nombre ASC
    """)
    productos = cur.fetchall()
    return render_template("cliente_productos.html", productos=productos)


# Agregar producto al carrito guardado en sesión
@app.route("/carrito/agregar/<int:id_producto>", methods=["POST"])
def agregar_carrito(id_producto):
    cantidad = int(request.form.get("cantidad", 1))
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM productos WHERE id_producto=%s", (id_producto,))
    producto = cur.fetchone()
    if not producto:
        return "Producto no encontrado", 404

    if "carrito" not in session:
        session["carrito"] = []

    carrito = session["carrito"]
    for item in carrito:
        if item["id_producto"] == id_producto:
            item["cantidad"] += cantidad
            break
    else:
        carrito.append({
            "id_producto": id_producto,
            "nombre": producto["nombre"],
            "precio": producto["precio"],
            "cantidad": cantidad
        })

    session["carrito"] = carrito
    flash("Producto agregado al carrito", "success")
    return redirect(url_for("cliente_productos"))


@app.route("/carrito", endpoint="cliente_carrito")
def ver_carrito():
    print("Ejecutando versión actualizada de ver_carrito")

    carrito = session.get("carrito", [])
    total = sum(item["precio"] * item["cantidad"] for item in carrito)

    cur = mysql.connection.cursor()

    cur.execute("""
        SELECT id_producto, nombre, precio, imagen
        FROM productos
        WHERE cod_categoria = (
            SELECT id_categoria FROM categorias WHERE nombre_categoria = 'Acompañamientos' LIMIT 1
        )
    """)
    acompanamientos = cur.fetchall()

    return render_template("cliente_carrito.html",
                           carrito=carrito,
                           total=total,
                           acompanamientos=acompanamientos)


# Confirmar pedido
@app.route("/pedido/confirmar", methods=["POST"])
def hacer_pedido():
    carrito = session.get("carrito", [])
    if not carrito:
        flash("El carrito está vacío", "warning")
        return redirect(url_for("cliente_productos"))

    total = sum(item["precio"] * item["cantidad"] for item in carrito)

    # Sumar acompañamientos seleccionados 
    acompanamientos_ids = request.form.getlist("acompanamientos")
    if acompanamientos_ids:
        cur = mysql.connection.cursor()
        formato = ','.join(['%s'] * len(acompanamientos_ids))
        cur.execute(f"SELECT SUM(precio) as total_acomp FROM productos WHERE id_producto IN ({formato})", acompanamientos_ids)
        total_acomp = cur.fetchone()["total_acomp"] or 0
        total += total_acomp

    tipo_entrega = request.form.get("tipo_entrega", "restaurante")
    metodo_pago = request.form.get("metodo_pago", "efectivo")

    # Si es domicilio, pedimos datos adicionales
    direccion = None
    telefono = None
    if tipo_entrega == "domicilio":
        direccion = request.form.get("direccion")
        telefono = request.form.get("telefono_envio")

    id_usuario = session.get("id_usuario", 1)  # temporal

    cur = mysql.connection.cursor()
    cur.execute("""
        INSERT INTO pedidos (
            cod_usuario, fecha, hora, total, estado,
            tipo_entrega, metodo_pago, direccion, telefono
        ) VALUES (
            %s, CURDATE(), CURTIME(), %s, 'pendiente',
            %s, %s, %s, %s
        )
    """, (
        id_usuario, total, tipo_entrega, metodo_pago,
        direccion, telefono
    ))
    id_pedido = cur.lastrowid

    for item in carrito:
        cur.execute("""
            INSERT INTO detalle_pedido (
                cod_pedido, cod_producto, cantidad, precio_unitario
            ) VALUES (%s, %s, %s, %s)
        """, (
            id_pedido, item["id_producto"],
            item["cantidad"], item["precio"]
        ))

    if acompanamientos_ids:
        for id_acomp in acompanamientos_ids:
            cur.execute("""
                INSERT INTO detalle_pedido (cod_pedido, cod_producto,
                cantidad, precio_unitario)
                SELECT %s, id_producto, 1, precio
                FROM productos WHERE id_producto = %s
            """, (id_pedido, id_acomp))

    mysql.connection.commit()
    session.pop("carrito", None)

    flash("Pedido registrado correctamente", "success")
    return redirect(url_for("cliente_dashboard"))


@app.route("/mis_pedidos")
def ver_pedidos():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT id_pedido, cod_usuario, fecha, hora, total, estado,
               tipo_entrega, metodo_pago, direccion, telefono
        FROM pedidos ORDER BY id_pedido DESC
    """)
    pedidos = cur.fetchall()
    return render_template("mis_pedidos.html", pedidos=pedidos)



# Eliminar producto del carrito
@app.route("/carrito/eliminar/<int:id_producto>")
def eliminar_carrito(id_producto):
    carrito = session.get("carrito", [])
    nuevo_carrito = [
        item for item in carrito if item["id_producto"] != id_producto
    ]
    session["carrito"] = nuevo_carrito
    flash("Producto eliminado del carrito", "success")
    return redirect(url_for("cliente_carrito"))


# ------------------ LOGOUT ------------------
# Limpia la sesión completamente


# @app.route("/logout")
# def logout():
#    session.clear()
#    return redirect(url_for("login"))

# ------------------ RUN ------------------
# Ejecuta la app en modo debug (útil durante desarrollo)



# -------------------- EMPLEADO ----------------------
# ===================== LISTAR PRODUCTOS Y CATEGORÍAS =====================


from flask import render_template, request, jsonify
import MySQLdb.cursors  # 👈 necesario para usar DictCursor

@app.route("/registrar_empleado")
def registrar_empleado():
    cur = mysql.connection.cursor()

    # 🔹 Obtener categorías
    cur.execute("SELECT id_categoria, nombre_categoria FROM categorias")
    categorias_raw = cur.fetchall()

    categorias = []
    for c in categorias_raw:
        if isinstance(c, dict):
            categorias.append({
                'id_categoria': c.get('id_categoria'),
                'nombre_categoria': c.get('nombre_categoria')
            })
        else:
            categorias.append({
                'id_categoria': c[0],
                'nombre_categoria': c[1]
            })

    # 🔹 Obtener productos (ahora incluye el campo estado)
    cur.execute("""
        SELECT 
            p.id_producto, 
            p.nombre, 
            p.precio, 
            p.cod_categoria AS id_categoria, 
            p.estado,  -- 👈 se agregó
            c.nombre_categoria
        FROM productos p
        LEFT JOIN categorias c ON p.cod_categoria = c.id_categoria
    """)
    productos_raw = cur.fetchall()

    productos = []
    for p in productos_raw:
        if isinstance(p, dict):
            productos.append({
                'id_producto': p.get('id_producto'),
                'nombre': p.get('nombre'),
                'precio': p.get('precio'),
                'id_categoria': p.get('id_categoria'),
                'nombre_categoria': p.get('nombre_categoria'),
                'estado': p.get('estado', 'Disponible')  # 👈 por si viene NULL
            })
        else:
            productos.append({
                'id_producto': p[0],
                'nombre': p[1],
                'precio': p[2],
                'id_categoria': p[3],
                'estado': p[4] if len(p) > 4 else 'Disponible',
                'nombre_categoria': p[5] if len(p) > 5 else ''
            })

    cur.close()

    # 🚀 Debug opcional (puedes quitarlo si todo anda bien)
    print("=== PRODUCTOS NORMALIZADOS ===")
    for p in productos:
        print(p)

    return render_template("registrar_empleado.html", productos=productos, categorias=categorias)


# 🔹 Actualizar estado del producto
@app.route("/actualizar_estado_producto", methods=["POST"])
def actualizar_estado_producto():
    data = request.get_json()
    id_producto = data.get("id_producto")
    nuevo_estado = data.get("estado")

    if not id_producto or not nuevo_estado:
        return jsonify({"success": False, "msg": "Datos incompletos"}), 400

    try:
        cur = mysql.connection.cursor()
        cur.execute("""
            UPDATE productos 
            SET estado = %s 
            WHERE id_producto = %s
        """, (nuevo_estado, id_producto))
        mysql.connection.commit()
        cur.close()
        return jsonify({"success": True, "nuevo_estado": nuevo_estado})
    except Exception as e:
        mysql.connection.rollback()
        return jsonify({"success": False, "msg": str(e)}), 500

# ==================== Registrar pagos ==============


from datetime import datetime
from flask import request, redirect, url_for, flash
import json

@app.route('/registrar_pago_restaurante', methods=['POST'])
def registrar_pago_restaurante():
    try:
        # Obtener datos del formulario
        id_mesa = request.form.get('id_mesa')
        total = float(request.form.get('total', 0))
        dinero_cliente = float(request.form.get('dinero_cliente', 0))
        productos_json = request.form.get('productos', '[]')
        productos = json.loads(productos_json)

        fecha = datetime.now().date()
        hora = datetime.now().strftime("%H:%M:%S")

        # Validación básica
        if not id_mesa or total <= 0 or len(productos) == 0:
            flash("Datos de la orden incompletos.", "error")
            return redirect(url_for('mesas_empleado'))

        cur = mysql.connection.cursor()

        # Insertar pago principal
        cur.execute("""
            INSERT INTO pagos_restaurante (id_mesa, fecha, hora, total)
            VALUES (%s, %s, %s, %s)
        """, (id_mesa, fecha, hora, total))
        id_pago_restaurante = cur.lastrowid

        # Insertar detalles de cada producto
        for p in productos:
            cur.execute("""
                INSERT INTO detalle_pedido_restaurante 
                (id_pago_restaurante, id_producto, cantidad, precio_unitario)
                VALUES (%s, %s, %s, %s)
            """, (
                id_pago_restaurante,
                p['id_producto'],
                p['cantidad'],
                p['precio']
            ))

        mysql.connection.commit()
        flash("✅ Pago registrado correctamente.", "success")
        return redirect(url_for('mesas_empleado'))

    except Exception as e:
        print("Error MySQL:", e)
        flash(f"Error al registrar el pago: {e}", "error")
        return redirect(url_for('mesas_empleado'))

    finally:
        cur.close()

# ==================== HISTORIAL DE PAGOS ====================
from flask import request, render_template
from datetime import datetime

@app.route("/historial_pagos_restaurante", methods=["GET"])
def historial_pagos_restaurante():
    cur = mysql.connection.cursor()

    # Parámetros de búsqueda
    query = request.args.get("query", "").strip()

    if query:
        cur.execute("""
            SELECT * FROM pagos_restaurante
            WHERE CAST(id_pago_restaurante AS CHAR) LIKE %s
               OR DATE_FORMAT(fecha, '%%Y-%%m-%%d') LIKE %s
               OR hora LIKE %s
            ORDER BY fecha DESC, hora DESC
        """, (f"%{query}%", f"%{query}%", f"%{query}%"))
    else:
        cur.execute("SELECT * FROM pagos_restaurante ORDER BY fecha DESC, hora DESC")

    pagos = cur.fetchall()

    historial = []
    for pago in pagos:
        cur.execute("""
            SELECT d.*, p.nombre
            FROM detalle_pedido_restaurante d
            JOIN productos p ON d.id_producto = p.id_producto
            WHERE d.id_pago_restaurante = %s
        """, (pago["id_pago_restaurante"],))
        detalles = cur.fetchall()

        pago["detalles"] = detalles
        historial.append(pago)

    cur.close()
    return render_template("historial_pagos_restaurante.html", historial=historial)


# ===================== MESAS Y ORDENES =====================
@app.route('/mesas_empleado')
def mesas_empleado():
    return render_template('mesas_empleado.html')


@app.route('/orden/<int:mesa_id>', methods=['GET', 'POST'])
def orden_mesa(mesa_id):
    cur = mysql.connection.cursor()
    
    if request.method == 'POST':
        productos_seleccionados = request.form.getlist('producto')
        total = request.form.get('total', 0)

        # Validación: Asegurarse de que hay productos seleccionados
        if not productos_seleccionados:
            flash("Por favor, selecciona al menos un producto para realizar el pedido.", "error")
            return redirect(url_for('orden_mesa', mesa_id=mesa_id))  # Redirige a la misma página si no hay productos

        # Validación: Si el total no es proporcionado, lo calculamos sumando los precios de los productos seleccionados
        if total == '0' or total == '':
            total_calculado = 0
            for producto_id in productos_seleccionados:
                cur.execute("SELECT precio FROM productos WHERE id = %s", (producto_id,))
                producto = cur.fetchone()
                if producto:
                    total_calculado += producto['precio']
            total = total_calculado

        # Insertar el pago en la tabla pagos_restaurante
        cur2 = mysql.connection.cursor()
        cur2.execute("""
            INSERT INTO pagos_restaurante (id_mesa, fecha, hora, total)
            VALUES (%s, CURDATE(), CURTIME(), %s)
        """, (mesa_id, total))
        mysql.connection.commit()
        cur2.close()
        cur.close()

        # Flash mensaje de éxito
        flash("Pedido realizado con éxito. El pago se ha registrado correctamente.", "success")
        
        # Redirigir a la página de mesas después del pago
        return redirect(url_for('mesas_empleado'))
    
    # Obtener categorías y productos si se usa GET
    cur.execute("SELECT * FROM categorias")
    categorias = cur.fetchall()

    cur.execute("SELECT * FROM productos")
    productos = cur.fetchall()

    cur.close()
    return render_template('calculadora.html', mesa=mesa_id, categorias=categorias, productos=productos)

    
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash
from flask_mysqldb import MySQL, MySQLdb
from datetime import datetime
from MySQLdb import IntegrityError 

# ===================== RESERVAS empleado =====================

@app.route("/agregar_reserva", methods=["POST"])
def agregar_reserva():
    fecha = request.form["fecha"]

    cur = mysql.connection.cursor()

    # Validar reserva por fecha
    cur.execute("SELECT COUNT(*) FROM reservas WHERE fecha = %s", (fecha,))
    result = cur.fetchone()
    count = result['COUNT(*)'] if result else 0

    if count > 0:
        flash("Ya existe una reserva para esta fecha. Solo se permite una por día.", "error")
        cur.close()
        return redirect(url_for("reservas_empleado"))

    nombre = request.form["nombre"]
    documento = request.form["documento"]
    telefono = request.form["telefono"]
    hora = request.form["hora"]
    cant_personas = request.form["cant_personas"]
    tipo_evento = request.form["tipo_evento"]
    comentarios = request.form.get("comentarios", "")
    id_usuario = request.form["id_usuario"]
    estado = request.form.get("estado", "disponible")

    # Validar que el id_usuario exista
    cur.execute("SELECT id_usuario FROM usuarios WHERE id_usuario = %s", (id_usuario,))
    usuario = cur.fetchone()
    if not usuario:
        flash("⚠️ El ID de usuario ingresado no existe. Por favor ingresa un ID válido.", "error")
        cur.close()
        return redirect(url_for("reservas_empleado"))

    try:
        cur.execute("""
            INSERT INTO reservas (
                nombre, documento, telefono, fecha, hora,
                cant_personas, tipo_evento, comentarios, id_usuario, estado
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            nombre, documento, telefono, fecha, hora,
            cant_personas, tipo_evento, comentarios, id_usuario, estado
        ))
        mysql.connection.commit()
        flash("Reserva agregada exitosamente.", "success")

    except IntegrityError as e:
        mysql.connection.rollback()
        flash(f"⚠️ Error de restricción de clave foránea: {str(e)}", "error")

    except Exception as e:
        mysql.connection.rollback()
        flash(f"Ocurrió un error inesperado: {str(e)}", "error")

    finally:
        cur.close()

    return redirect(url_for("reservas_empleado"))


@app.route("/editar_reserva/<int:id_reserva>", methods=["POST"])
def editar_reserva(id_reserva):
    nueva_fecha = request.form["fecha"]

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("""
        SELECT COUNT(*) AS total FROM reservas
        WHERE fecha = %s AND id_reserva != %s
    """, (nueva_fecha, id_reserva))
    count = cur.fetchone()
    count = count["total"] if count else 0

    if count > 0:
        flash("Ya existe una reserva para esta fecha. Solo se permite una por día.", "error")
        cur.close()
        return redirect(url_for("reservas_empleado"))

    # Actualizar datos
    nombre = request.form["nombre"]
    documento = request.form["documento"]
    telefono = request.form["telefono"]
    hora = request.form["hora"]
    cant_personas = request.form["cant_personas"]
    tipo_evento = request.form["tipo_evento"]
    comentarios = request.form["comentarios"]
    id_usuario = request.form["id_usuario"]
    estado = request.form.get("estado", "disponible")

    cur.execute("""
        UPDATE reservas SET
            nombre=%s,
            documento=%s,
            telefono=%s,
            fecha=%s,
            hora=%s,
            cant_personas=%s,
            tipo_evento=%s,
            comentarios=%s,
            id_usuario=%s,
            estado=%s
        WHERE id_reserva=%s
    """, (nombre, documento, telefono, nueva_fecha, hora,
          cant_personas, tipo_evento, comentarios, id_usuario, estado, id_reserva))
    mysql.connection.commit()
    cur.close()

    flash("Reserva editada exitosamente.", "success")
    return redirect(url_for("reservas_empleado"))



@app.route('/eliminar_reserva/<int:id_reserva>', methods=['POST'])
def eliminar_reserva(id_reserva):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("DELETE FROM reservas WHERE id_reserva = %s", (id_reserva,))
    mysql.connection.commit()
    cur.close()
    flash("Reserva eliminada correctamente.", "success")
    return redirect(url_for('reservas_empleado'))



@app.route("/reservas_empleado")
def reservas_empleado():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("""
        SELECT * FROM reservas
        WHERE estado IN ('Pendiente', 'Confirmada')
        ORDER BY fecha DESC, hora DESC
    """)
    reservas = cur.fetchall()
    cur.close()

    today = datetime.today().strftime('%Y-%m-%d')
    return render_template("reservas_empleado.html", reservas=reservas, today=today)


# ===================== CAMBIAR ESTADO DE RESERVA (EMPLEADO) =====================
from flask import jsonify

@app.route("/cambiar_estado_reserva_em/<int:id_reserva>", methods=["POST"])
def cambiar_estado_reserva_em(id_reserva):
    nuevo_estado = request.form.get("nuevo_estado")

    if nuevo_estado not in ["Confirmada", "Completada"]:
        return jsonify({"error": "Estado no válido"}), 400

    try:
        cur = mysql.connection.cursor()
        cur.execute("UPDATE reservas SET estado = %s WHERE id_reserva = %s", (nuevo_estado, id_reserva))
        mysql.connection.commit()
        cur.close()
        return jsonify({"message": f"Estado actualizado a {nuevo_estado}."}), 200
    except Exception as e:
        print(f"Error al actualizar estado: {e}")
        return jsonify({"error": "Error interno del servidor"}), 500



@app.route("/buscar_reservas", methods=["GET"])
def buscar_reservas():
    search_query = request.args.get("search_query", "").strip()

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    if search_query:
        like_query = f"%{search_query}%"
        cur.execute("""
            SELECT * FROM reservas
            WHERE estado IN ('Pendiente', 'Confirmada')
            AND (nombre LIKE %s OR documento LIKE %s OR telefono LIKE %s)
            ORDER BY fecha, hora
        """, (like_query, like_query, like_query))
    else:
        cur.execute("""
            SELECT * FROM reservas
            WHERE estado IN ('Pendiente', 'Confirmada')
            ORDER BY fecha, hora
        """)

    reservas = cur.fetchall()
    cur.close()

    today = datetime.today().strftime('%Y-%m-%d')
    return render_template("reservas_empleado.html", reservas=reservas, today=today)


@app.route("/historial_reservas_em")
def historial_reservas_em():
    query = request.args.get("query", "").strip()
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if query:
        like_query = f"%{query}%"
        cur.execute("""
            SELECT * FROM reservas
            WHERE estado = 'Completada' AND (
                nombre LIKE %s OR telefono LIKE %s OR documento LIKE %s
            )
            ORDER BY fecha DESC, hora DESC
        """, (like_query, like_query, like_query))
    else:
        cur.execute("""
            SELECT * FROM reservas
            WHERE estado = 'Completada'
            ORDER BY fecha DESC, hora DESC
        """)

    historial = cur.fetchall()
    cur.close()
    return render_template("historial_reservas_em.html", historial=historial)

# ===================== ORDENES REGISTRADAS =====================

@app.route('/ordenes_empleado')
def ordenes_empleado():
    cur = mysql.connection.cursor(DictCursor)
    cur.execute("SELECT * FROM pedidos ORDER BY fecha, hora")
    ordenes = cur.fetchall()
    cur.close()

    print("ORDENES para empleado:", ordenes)  # <--- para ver en consola

    return render_template("ordenes_empleado.html", ordenes=ordenes)


@app.route("/eliminar_orden/<int:id_pedido>")
def eliminar_orden(id_pedido):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM pedidos WHERE id_pedido = %s", (id_pedido,))
    mysql.connection.commit()
    cur.close()
    return redirect(url_for("ordenes_empleado"))


@app.route('/actualizar_estado/<int:id_pedido>', methods=['POST'])
def actualizar_estado(id_pedido):
    data = request.get_json()
    nuevo_estado = data.get("estado")

    cursor = mysql.connection.cursor()
    cursor.execute(
        "UPDATE pedidos SET estado=%s WHERE id_pedido=%s",
        (nuevo_estado, id_pedido)
    )
    mysql.connection.commit()
    cursor.close()

    return jsonify({
        "success": True,
        "id_pedido": id_pedido,
        "nuevo_estado": nuevo_estado
    })

# ===================== API JSON =====================


@app.route("/get_categorias")
def get_categorias():
    cur = mysql.connection.cursor()
    cur.execute("SELECT id_categoria, nombre_categoria FROM categorias")
    categorias = cur.fetchall()
    cur.close()
    return jsonify(categorias)


@app.route("/get_productos/<int:id_categoria>")
def get_productos(id_categoria):
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT id_producto, nombre, precio
        FROM productos
        WHERE id_categoria = %s
    """, (id_categoria,))
    productos = cur.fetchall()
    cur.close()
    return jsonify(productos)


# ===================== TEST DB =====================
@app.route("/test_db")
def test_db():
    try:
        cur = mysql.connection.cursor()
        cur.execute("SHOW TABLES;")
        tablas = cur.fetchall()
        cur.close()
        return f"✅ Conectado correctamente. Tablas: {tablas}"
    except Exception as e:
        return f"❌ Error de conexión: {e}"



# ===================== MAIN =====================
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)
